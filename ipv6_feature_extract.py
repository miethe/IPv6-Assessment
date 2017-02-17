##################################
# Author: Nick Miethe
# Date: 1/9/2017
# Requirements: This program runs on Python 2.7 and will require openpyxl 2.4.1 or greater
# Openpyxl can be obtained by running the command $ pip install openpyxl
#
# This script takes the output of ipv6_feature_extract.py and creates an excel file of tabs with each
# tab being a different device type. In each tab are the various devices and their responding
# IPv6 features, color coded based on status.
###################################


from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED, GREEN, BLUE
from openpyxl.utils import coordinate_from_string, column_index_from_string


Tk().withdraw()
print "Open IPv6 Assessment Code"
ip_assess = askopenfilename()

#load in workbook from filename and get active worksheet
ip_wb = load_workbook(ip_assess)
ip_ws = ip_wb.active

#create workbook and sheet for output
out_wb = Workbook()
out_ws = out_wb.active

bDevices = True
row_int = 1
row_max = ip_ws.max_row
col_max = ip_ws.max_column
new_row = 1

while new_row < row_max:
	if ip_ws.cell(row=new_row,column=1).value and "Group:" in ip_ws.cell(row=new_row,column=1).value:
		out_ws = out_wb.create_sheet(str(ip_ws.cell(row=new_row,column=1).value)[7:].replace(":", "").replace("*",""))
		row_int = 1
	#if ready for new device
	if bDevices is True:
		if ip_ws.cell(row=new_row,column=2).value:
			#If starting new device
			if "Device Name" in ip_ws.cell(row=new_row,column=2).value:
				#print a blank line
				out_ws.cell(row=row_int,column=1).value=""
				#move to next row and 2nd next row
				row_int +=1
				new_row +=2
				#pull OS from file
				OS = str(ip_ws.cell(row=new_row,column=2).value)+" "+str(ip_ws.cell(row=new_row,column=3).value)+" "+str(ip_ws.cell(row=new_row,column=6).value)
				#write OS to output
				out_ws.cell(row=row_int,column=1).value=OS
				#Bold font
				out_ws.cell(row=row_int,column=1).font = Font(bold=True) 
				row_int +=1
				#begin parsing capabilities
				bDevices = False
	else:
		if ip_ws.cell(row=new_row,column=1).value:
			if "OS Version Capability Status" in ip_ws.cell(row=new_row,column=1).value:
				#move down 4 rows after "OS Version Capability Status"
				new_row +=4
				#while row isn't blank or a delimiter (end of section)
				while ip_ws.cell(row=new_row,column=1).value is None or "-" not in ip_ws.cell(row=new_row,column=1).value:
					out_ws.cell(row=row_int,column=1).value = ip_ws.cell(row=new_row, column=4).value
					#If capability is N, Not Capable
					if 'N' in ip_ws.cell(row=new_row,column=2).value:
						#change font to Green
						out_ws.cell(row=row_int,column = 1).font = Font(color=RED)
						out_ws.cell(row=row_int,column=2).value = "Not Capable"
					#If capability is C, Capable
					elif 'C' in ip_ws.cell(row=new_row,column=2).value:
						#change font to Green
						out_ws.cell(row=row_int,column = 1).font = Font(color=GREEN)
						out_ws.cell(row=row_int,column=2).value = "Capable"
					#If capability is U, Upgrade
					elif 'U' in ip_ws.cell(row=new_row,column=2).value:
						#change font to Blue
						out_ws.cell(row=row_int,column = 1).font = Font(color=BLUE)
						out_ws.cell(row=row_int,column=2).value = "Upgrade"
					#move to next rows
					row_int+=1
					new_row+=1
				#get new device
				bDevices = True
	new_row+=1
out_wb.save("Extracted Features.xlsx")