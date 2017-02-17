from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl import Workbook
from openpyxl import load_workbook

Tk().withdraw()
print "Open Exported IPv6 Feature Mapping"
file_path = askopenfilename()

#load in workbook from filename and get active worksheet
feat_wb = load_workbook(file_path)
feat_ws = feat_wb.active

print "Open Report"
file_path = askopenfilename()

#create workbook and sheet for output
ip_wb = Workbook(file_path)
ip_ws = ip_wb.active

#setup variables for report
report_row_int = 1
row_max = ip_ws.max_row
col_max = ip_ws.max_column

#setup variables for feature mapping
feat_row_int = 1
feat_row_max = feat_ws.max_row
feat_col_max = feat_ws.max_column

while report_row_int < row_max:
	#if comment section not blank
	if ip_ws.cell(row=report_row_int,column=7).value:
		feat_str = ip_ws.cell(row=report_row_int,column=7).value
		#get all features from cell
		feat_list = feat_str.split('/n')
		for feat in feat_list:
			#check all features from web page to match against feature in report
			while feat_row_int < feat_row_max:
				if feat in feat_ws.cell(row=feat_row_int,column=1).value or feat in feat_ws.cell(row=feat_row_int,column=2).value:
					#check if OS type
				else:
					feat_row_int += 1
			feat_row_int = 1
		#next line
		report_row_int += 1
	else:
		#skip to next line
		report_row_int + = 1