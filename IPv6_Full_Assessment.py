####################
# Nick Miethe
# 2/3/2017
# Cisco Systems
#
# Requirements: This program runs on Python 2.7 and will require openpyxl 2.4.1 or greater
# Libraries: Openpyxl, Tkinter
#
# This program accepts in a text file of IPv6 Assessment. It will then create an excel file
# with one tab per group type. Within each tab will be all devices with the IPv6 features listed
# and color-coded for capability.
#
# The program also supports filling out an IPv6 report, though is still a WIP.
#
####################

from Tkinter import Tk
from tkFileDialog import askopenfilename
from os import listdir
from os.path import isfile, join
import re
import glob
import csv

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED, GREEN, BLUE
from openpyxl.utils import coordinate_from_string, column_index_from_string


def Convert_xl():

	print "Open Text file"
	textfile = askopenfilename()

	spamReader = csv.reader((open(textfile, 'rb')), delimiter='|',quotechar='"')

	name_list = textfile.split('/')
	filename = name_list[len(name_list)-1].split('.')[0]
	wb = Workbook()
	sheet = wb.active
	newName = textfile
	for rowx, row in enumerate(spamReader):
  		for colx, value in enumerate(row):
  			sheet.cell(row=rowx+1,column=colx+1).value=value
  			#sheet.write(rowx, colx, value)
  	return wb

def Extract_features(ip_wb):
	
	ip_ws = ip_wb.active

	#create workbook and sheet for output
	out_wb = Workbook()
	out_ws = out_wb.active

	bDevices = True
	row_int = 1
	row_max = ip_ws.max_row
	col_max = ip_ws.max_column
	new_row = 1

	while new_row < row_max:
		if ip_ws.cell(row=new_row,column=1).value and "Group:" in ip_ws.cell(row=new_row,column=1).value and \
			"End" not in ip_ws.cell(row=new_row,column=1).value:
			out_ws = out_wb.create_sheet(str(ip_ws.cell(row=new_row,column=1).value)[7:].replace(":", "").replace("*",""))
			row_int = 1
		#if ready for new device
		if bDevices is True:
			if ip_ws.cell(row=new_row,column=2).value:
				#If starting new device
				if "Device Name" in ip_ws.cell(row=new_row,column=2).value:
					#print a blank line
					out_ws.cell(row=row_int,column=1).value=""
					#move to next row and 2nd next row
					row_int +=1
					new_row +=2
					#pull OS from file
					OS = str(ip_ws.cell(row=new_row,column=2).value)+" "+str(ip_ws.cell(row=new_row,column=3).value)+" "+str(ip_ws.cell(row=new_row,column=6).value)
					#write OS to output
					out_ws.cell(row=row_int,column=1).value=OS
					#Bold font
					out_ws.cell(row=row_int,column=1).font = Font(bold=True) 
					row_int +=1
					#begin parsing capabilities
					bDevices = False
		else:
			if ip_ws.cell(row=new_row,column=1).value:
				if "OS Version Capability Status" in ip_ws.cell(row=new_row,column=1).value:
					#move down 4 rows after "OS Version Capability Status"
					new_row +=4
					#while row isn't blank or a delimiter (end of section)
					while ip_ws.cell(row=new_row,column=1).value is None or "-" not in ip_ws.cell(row=new_row,column=1).value:
						out_ws.cell(row=row_int,column=1).value = ip_ws.cell(row=new_row, column=4).value
						#If capability is N, Not Capable
						if 'N' in ip_ws.cell(row=new_row,column=2).value:
							#change font to Green
							out_ws.cell(row=row_int,column = 1).font = Font(color=RED)
							out_ws.cell(row=row_int,column=2).value = "Not Capable"
						#If capability is C, Capable
						elif 'C' in ip_ws.cell(row=new_row,column=2).value:
							#change font to Green
							out_ws.cell(row=row_int,column = 1).font = Font(color=GREEN)
							out_ws.cell(row=row_int,column=2).value = "Capable"
						#If capability is U, Upgrade
						elif 'U' in ip_ws.cell(row=new_row,column=2).value:
							#change font to Blue
							out_ws.cell(row=row_int,column = 1).font = Font(color=BLUE)
							out_ws.cell(row=row_int,column=2).value = "Upgrade"
						#move to next rows
						row_int+=1
						new_row+=1
					#get new device
					bDevices = True
		new_row+=1
	return out_wb
	
def Build_report(ip_wb):

	ip_ws_names = ip_wb.sheetnames

	ws_up = ""
	ws_cap = ""
	ws_other = ""
	ws_not = ""

	for ws in ip_ws_names:
		if "Upgrade" in ws:
			ws_up = ws
		elif "Capable" in ws and "Not" not in ws:
			ws_cap = ws
		elif "Not" in ws:
			ws_not = ws
		else:
			ws_other = ws

	ip_ws = ip_wb.get_sheet_by_name(ws_other)

	bDevices = True
	row_int = 1
	row_max = ip_ws.max_row
	col_max = ip_ws.max_column
	new_row = 1
	sDeviceName = ""

	while new_row < row_max:
		if ip_ws.cell(row=new_row,column=1).value is not None and \
			"Pv6" not in ip_ws.cell(row=new_row, column=1).value and \
			ip_ws.cell(row=new_row, column=1).value is not "":
		
			sDevice = ip_ws.cell(row=new_row, column=1).value
			print sDevice
			sDeviceName = sDevice.split()[0]
			sProductType = sDevice.split()[1]
			sOS = sDevice.split()[2]
			bDevices = True
	
		if bDevices is True:
			while ip_ws.cell(row=new_row, column=1).value is not None and \
				ip_ws.cell(row=new_row, column=1).value is not "":
			
				if ip_ws.cell(row=new_row, column=2).value and \
					"Capable" in ip_ws.cell(row=new_row, column=2).value and \
					"Not" not in ip_ws.cell(row=new_row, column=2).value and \
					ws_cap != "":
				
					sFeature = ip_ws.cell(row=new_row, column=1).value
					ip_ws = ip_wb.get_sheet_by_name(ws_cap)
					new_row_max = ip_ws.max_row
					n_new_row = 1
					while n_new_row <= new_row_max:
						if ip_ws.cell(row=n_new_row, column=1).value == sDeviceName:
							if ip_ws.cell(row=n_new_row, column=8).value:
								ip_ws.cell(row=n_new_row, column=8).value = \
									ip_ws.cell(row=n_new_row, column=8).value + "\n" + sFeature
							else:
								ip_ws.cell(row=n_new_row, column=8).value = sFeature
							n_new_row = new_row_max+1
						else:
							n_new_row +=1
					new_row +=1
				
				elif ip_ws.cell(row=new_row, column=2).value and \
					"Upgrade" in ip_ws.cell(row=new_row, column=2).value and \
					ws_up != "":
				
					sFeature = ip_ws.cell(row=new_row, column=1).value
					ip_ws = ip_wb.get_sheet_by_name(ws_up)
					new_row_max = ip_ws.max_row
					n_new_row = 1
					while n_new_row <= new_row_max:
						if ip_ws.cell(row=n_new_row, column=1).value == sDeviceName:
							if ip_ws.cell(row=n_new_row, column=8).value:
								ip_ws.cell(row=n_new_row, column=8).value = \
									ip_ws.cell(row=n_new_row, column=8).value + "\n" + sFeature
							else:
								ip_ws.cell(row=n_new_row, column=8).value = sFeature
							n_new_row = new_row_max+1
						else:
							n_new_row +=1
					new_row +=1
				
				else: 
					new_row +=1
				ip_ws = ip_wb.get_sheet_by_name(ws_other)
			bDevices = False
		new_row+=1
		
	return ip_wb
	
def main():
	
	workbook = Convert_xl()
	workbook = Extract_features(workbook)
	
	workbook.save("Report_Complete.xlsx")
	
if __name__ == "__main__":
    main()