############################# 
# Author: Nick Miethe
# Date: 1/11/2017
# Requirements: This program runs on Python 2.7 and will require openpyxl 2.4.1 or greater
# Openpyxl can be obtained by running the command $ pip install openpyxl
#
# This script takes the output of ipv6_feature_extract.py which has been imported into an IPv6 report
# and fills out the info on the relevant tabs.
#############################


from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl import Workbook
from openpyxl import load_workbook

Tk().withdraw()
print "Open Extracted Features"
ip_assess = askopenfilename()

ip_wb = load_workbook(ip_assess)
ip_ws_names = ip_wb.sheetnames

ws_up = ""
ws_cap = ""
ws_other = ""
ws_not = ""

for ws in ip_ws_names:
	if "Upgrade" in ws:
		ws_up = ws
	elif "Capable" in ws and "Not" not in ws:
		ws_cap = ws
	elif "Not" in ws:
		ws_not = ws
	else:
		ws_other = ws

ip_ws = ip_wb.get_sheet_by_name(ws_other)

bDevices = True
row_int = 1
row_max = ip_ws.max_row
col_max = ip_ws.max_column
new_row = 1
sDeviceName = ""

while new_row < row_max:
	if ip_ws.cell(row=new_row,column=1).value is not None and \
		"Pv6" not in ip_ws.cell(row=new_row, column=1).value and \
		ip_ws.cell(row=new_row, column=1).value is not "":
		
		sDevice = ip_ws.cell(row=new_row, column=1).value
		sDeviceName = sDevice.split()[0]
		sProductType = sDevice.split()[1]
		sOS = sDevice.split()[2]
		bDevices = True
	
	if bDevices is True:
		while ip_ws.cell(row=new_row, column=1).value is not None and \
			ip_ws.cell(row=new_row, column=1).value is not "":
			
			if ip_ws.cell(row=new_row, column=2).value and \
				"Capable" in ip_ws.cell(row=new_row, column=2).value and \
				"Not" not in ip_ws.cell(row=new_row, column=2).value and \
				ws_cap != "":
				
				sFeature = ip_ws.cell(row=new_row, column=1).value
				ip_ws = ip_wb.get_sheet_by_name(ws_cap)
				new_row_max = ip_ws.max_row
				n_new_row = 1
				while n_new_row <= new_row_max:
					if ip_ws.cell(row=n_new_row, column=1).value == sDeviceName:
						if ip_ws.cell(row=n_new_row, column=8).value:
							ip_ws.cell(row=n_new_row, column=8).value = \
								ip_ws.cell(row=n_new_row, column=8).value + "\n" + sFeature
						else:
							ip_ws.cell(row=n_new_row, column=8).value = sFeature
						n_new_row = new_row_max+1
					else:
						n_new_row +=1
				new_row +=1
				
			elif ip_ws.cell(row=new_row, column=2).value and \
				"Upgrade" in ip_ws.cell(row=new_row, column=2).value and \
				ws_up != "":
				
				sFeature = ip_ws.cell(row=new_row, column=1).value
				ip_ws = ip_wb.get_sheet_by_name(ws_up)
				new_row_max = ip_ws.max_row
				n_new_row = 1
				while n_new_row <= new_row_max:
					if ip_ws.cell(row=n_new_row, column=1).value == sDeviceName:
						if ip_ws.cell(row=n_new_row, column=8).value:
							ip_ws.cell(row=n_new_row, column=8).value = \
								ip_ws.cell(row=n_new_row, column=8).value + "\n" + sFeature
						else:
							ip_ws.cell(row=n_new_row, column=8).value = sFeature
						n_new_row = new_row_max+1
					else:
						n_new_row +=1
				new_row +=1
				
			else: 
				new_row +=1
			ip_ws = ip_wb.get_sheet_by_name(ws_other)
		bDevices = False
	new_row+=1
ip_wb.save(ip_assess.split(".")[0] + "_Complete.xlsx")